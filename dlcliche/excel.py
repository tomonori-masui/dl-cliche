from dlcliche.utils import *
import openpyxl as opx
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from copy import copy
import re

def opx_copy_cell(source_cell, target_cell, style_copy=False):
    """openpyxl helper: Copy cell from source to target.
    Cells can be on different sheets.

    Args:
        style_copy: Flag to copy style or not.
    """
    target_cell._value = source_cell._value
    target_cell.data_type = source_cell.data_type
    if style_copy and source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    if source_cell.hyperlink:
        target_cell._hyperlink = copy(source_cell.hyperlink)
    if source_cell.comment:
        target_cell.comment = copy(source_cell.comment)

def opx_copy_cell_style(source_cell, target_cell):
    """openpyxl helper: Copy cell style only from source to target.
    """
    target_cell._style = copy(source_cell._style)

def opx_copy_row(sh_src, row_src, sh_dest, row_dest, n_col=None, style_copy=False, debug=False):
    """openpyxl helper: Copy row.

    Args:
        sh_src: Source sheet.
        row_src: Source row number, range is as usual: `[0, max-1]`.
        sh_dest: Destination sheet.
        row_dest: Destination row number, same range as source.
        n_col: Number of columns to copy, n_col=5 will copy five sequential columns.
        style_copy: Flag to copy style or not.
    """
    if n_col is None: n_col = sh_src.max_column
    if debug:
        print('<', 1, row_src+1, sh_src.cell(column=1, row=row_src+1).value)
        print('>', 1, row_dest+1, sh_dest.cell(column=1, row=row_dest+1).value)
    for c in range(n_col):
        opx_copy_cell(sh_src.cell(column=c+1, row=row_src+1), sh_dest.cell(column=c+1, row=row_dest+1),
                     style_copy=style_copy)

def opx_duplicate_style(sh, row_src, row_dest, n_row=None, debug=False):
    """openpyxl helper: Copy style among rows.

    Args:
        sh: Sheet.
        row_src: Source row number, range is as usual: `[0, max-1]`.
        row_dest: Destination row number, same range as source.
        n_row: Number of rows to copy, n_col=5 will copy five sequential rows.
        style_copy: Flag to copy style or not.
    """
    n_col = sh.max_column
    if n_row is None:
        n_row = sh.max_row - row_dest
    if debug:
        print('duplicate', row_src, 'style to', row_dest, '-', row_dest+n_row - 1)
    for r in range(row_dest, row_dest+n_row):
        for c in range(n_col):
            opx_copy_cell_style(sh.cell(column=c+1, row=row_src+1), sh.cell(column=c+1, row=r+1))

opx_ILLEGAL_CHARACTERS_RE = re.compile(
    r'[\000-\010]|[\013-\014]|[\016-\037]|[\x00-\x1f\x7f-\x9f]|[\uffff]')
def opx_remove_illegal_char(data):
    """Remove ILLEGAL CHARACTERS.
    Thanks to https://qiita.com/analytics-hiro/items/e6581019a11de798002b
    """
    if isinstance(data, str):
        return opx_ILLEGAL_CHARACTERS_RE.sub("", data)
    else:
        return data

def opx_df_to_ws(workbook, sheetname, df, start_row=1, start_col=1, index=True, header=True, index_filter=None):
    """Write all data in a DataFrame to Excel worksheet.
    
    Args:
        workbook: Target workbook object.
        sheetname: Target worksheet name, will be created if not there.
        df: DataFrame object to read from.
        start_row: Target worksheet row to start writing.
        start_col: Target worksheet column to start writing.
        index: Put index data also if this is True.
        header: Put columns as title row on top of data if this is True.
        index_filter: Function to filter index value like `lambda x: x[:5]` to limit up to 5 letters.

    Returns:
        n_row: Number of rows written, regardless of starting position.
        n_col: Number of columns in the same fashion.
    """
    # Get existing or new worksheet
    if sheetname in workbook.sheetnames:
        ws = workbook[sheetname]
    else:
        ws = workbook.create_sheet(title=sheetname)

    # Put header manually, avoiding two lines...
    n_row, n_col = 0, 0
    if header:
        ws.cell(row=start_row, column=start_col, value=df.index.name)
        for ci, value in enumerate(df.columns, start_col+1):
            ws.cell(row=start_row, column=ci, value=opx_remove_illegal_char(value))
        start_row += 1
        n_row += 1
    # Write all data into the sheet
    for ri, row in enumerate(dataframe_to_rows(df, index=index, header=False), start_row):
        if header:
            if ri == start_row:
                continue # Header have been done already.
            else:
                ri -= 1
        for ci, value in enumerate(row, start_col):
            if index_filter is not None and ci == start_col:
                value = index_filter(value)
            ws.cell(row=ri, column=ci, value=opx_remove_illegal_char(value))
            if n_col < ci+1: n_col = ci+1
        if n_row < ri: n_row = ri
    return n_row, n_col

from openpyxl.utils import get_column_letter
def opx_bar_chart(dest_ws, RC, src_ws, TL, BR, figtitle='', xtitle='', ytitle='', figsize=None):
    """Draw bar chart of source data placed in a rectangle area.
    Leftmost column is index (= category).
    
    Args:
        dest_ws: Worksheet to append bar chart.
        RC: Row and column position to append bar chart.
        src_ws: Worksheet where chart data are.
        TL: Top left cell `(row, column)` of data area.
        BR: Bottom right cell of data area.
        figtitle: Chart title.
        xtitle: X title.
        ytitle: Y title.
        figsize: Size `(width, height)` of chart.
    """
    chart = opx.chart.BarChart()
    chart.type = 'col'
    chart.style = 10
    chart.title = figtitle
    chart.y_axis.title = ytitle
    chart.x_axis.title = xtitle

    cats = opx.chart.Reference(src_ws, min_col=TL[1], min_row=TL[0], max_col=TL[1], max_row=BR[0])
    data = opx.chart.Reference(src_ws, min_col=TL[1]+1, min_row=TL[0], max_col=BR[1], max_row=BR[0])
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    if figsize is not None:
        chart.width = figsize[0]
        chart.height = figsize[1]
    
    chart.legend = None

    anchor = get_column_letter(RC[1]) + str(RC[0])
    dest_ws.add_chart(chart, anchor)

def opx_auto_adjust_column_width(worksheet, max_width=200, default_width=8, scaling=1.1, dont_narrower=False):
    """Auto adjust all column width in a worksheet.
    
    Args:
        worksheet: Worksheet object to adjust in place.
        max_width: Max width to limit, or setting None will not limit.
        default_width: Default width, used only when dont_narrower is True.
        scaling: Scaling factor to calculate width; `unicode_visible_width(unistr) * scaling`.
        dont_narrower: Set True if you don't want make columns get narrower.
    """
    column_widths = []
    for row in worksheet.iter_rows():
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], unicode_visible_width(str(cell.value)))
            except IndexError:
                column_widths.append(unicode_visible_width(str(cell.value)))

    for i, column_width in enumerate(column_widths):
        if dont_narrower:
            # https://groups.google.com/forum/#!topic/openpyxl-users/ROYUQyH50ro
            cur_width = worksheet.column_dimensions[get_column_letter(i + 1)].width
            if cur_width is None: cur_width = default_width
            column_width = cur_width if column_width < cur_width else column_width
        if max_width is not None:
            column_width = max_width if max_width < column_width else column_width
        worksheet.column_dimensions[get_column_letter(i + 1)].width = column_width * scaling
